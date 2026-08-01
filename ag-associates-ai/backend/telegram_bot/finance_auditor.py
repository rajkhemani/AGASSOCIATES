"""Finance Auditor — analyzes Excel financial statements & bank accounts.

Detects document type, extracts transactions, runs audit checks,
and returns a structured report.
"""

import re
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ── Column detection ─────────────────────────────────────────────────────

_DATE_KEYWORDS = ["date", "transaction date", "value date", "txn date", "posting date"]
_DESC_KEYWORDS = ["description", "narration", "particulars", "details", "transaction", "remarks", "narrative"]
_DEBIT_KEYWORDS = ["debit", "dr", "withdrawal", "withdrawn", "paid", "issue", "debit amount"]
_CREDIT_KEYWORDS = ["credit", "cr", "deposit", "deposited", "received", "credit amount"]
_BALANCE_KEYWORDS = ["balance", "running balance", "closing balance", "available balance", "bal"]

_RATIO_THRESHOLDS = {
    "current_ratio": {"min": 1.2, "max": 3.0},
    "debt_equity": {"min": 0.0, "max": 2.0},
    "profit_margin": {"min": 0.0},  # no max
}


def _find_col(headers: List[str], keywords: List[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        hl = str(h).strip().lower()
        for kw in keywords:
            if kw in hl:
                return i
    return None


def _parse_number(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    if s.startswith("₹") or s.startswith("$") or s.startswith("€") or s.startswith("£"):
        s = s[1:]
    try:
        return float(s)
    except ValueError:
        return None


# ── Document type detection ──────────────────────────────────────────────

def _detect_doc_type(headers: List[str], sample_rows: List[List]) -> str:
    hl = " ".join(str(h).lower() for h in headers)

    score_bank = 0
    score_bs = 0
    score_pl = 0

    if _find_col(headers, _DATE_KEYWORDS) is not None:
        score_bank += 2
    if _find_col(headers, _DESC_KEYWORDS) is not None:
        score_bank += 2
    if _find_col(headers, _DEBIT_KEYWORDS) is not None:
        score_bank += 2
    if _find_col(headers, _CREDIT_KEYWORDS) is not None:
        score_bank += 2
    if _find_col(headers, _BALANCE_KEYWORDS) is not None:
        score_bank += 1

    bs_terms = ["balance sheet", "assets", "liabilities", "equity", "property", "plant",
                 "fixed asset", "current asset", "current liability", "shareholder"]
    for t in bs_terms:
        if t in hl:
            score_bs += 1

    pl_terms = ["profit", "loss", "income", "revenue", "expense", "statement", "p&l",
                "turnover", "cost", "gross", "net profit"]
    for t in pl_terms:
        if t in hl:
            score_pl += 1

    if score_bank >= 4:
        return "bank_statement"
    if score_bs >= score_pl and score_bs >= 2:
        return "balance_sheet"
    if score_pl >= score_bs and score_pl >= 2:
        return "profit_loss"
    return "general"


# ── Bank statement auditor ───────────────────────────────────────────────

def _audit_bank_statement(headers: List[str], rows: List[List]) -> Dict:
    date_col = _find_col(headers, _DATE_KEYWORDS)
    desc_col = _find_col(headers, _DESC_KEYWORDS)
    debit_col = _find_col(headers, _DEBIT_KEYWORDS)
    credit_col = _find_col(headers, _CREDIT_KEYWORDS)
    bal_col = _find_col(headers, _BALANCE_KEYWORDS)

    transactions = []
    total_debit = 0.0
    total_credit = 0.0
    min_balance = float("inf")
    max_balance = float("-inf")
    date_range: Tuple[Optional[str], Optional[str]] = (None, None)
    anomalies = []

    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        txn = {}

        if date_col is not None:
            txn["date"] = str(row[date_col]).strip() if date_col < len(row) else ""
        if desc_col is not None:
            txn["desc"] = str(row[desc_col]).strip() if desc_col < len(row) else ""

        debit = _parse_number(row[debit_col]) if debit_col is not None and debit_col < len(row) else None
        credit = _parse_number(row[credit_col]) if credit_col is not None and credit_col < len(row) else None
        balance = _parse_number(row[bal_col]) if bal_col is not None and bal_col < len(row) else None

        if debit is not None and debit > 0:
            txn["debit"] = debit
            total_debit += debit
        if credit is not None and credit > 0:
            txn["credit"] = credit
            total_credit += credit

        if balance is not None:
            txn["balance"] = balance
            if balance < min_balance:
                min_balance = balance
            if balance > max_balance:
                max_balance = balance

        if txn.get("date") and txn.get("desc"):
            transactions.append(txn)

    # Running balance check
    if bal_col is not None and len(transactions) > 1:
        for i in range(1, len(transactions)):
            expected = (transactions[i - 1].get("balance", 0)
                        - transactions[i].get("debit", 0)
                        + transactions[i].get("credit", 0))
            actual = transactions[i].get("balance", 0)
            if expected and actual and abs(expected - actual) > 0.5:
                anomalies.append(f"Balance mismatch on row {i + 2}: expected {expected:.2f}, got {actual:.2f}")

    # Date range
    dates = [t.get("date", "") for t in transactions if t.get("date")]
    if dates:
        date_range = (min(dates), max(dates))

    # Large transactions
    large_threshold = max(total_credit, total_debit) * 0.1 if max(total_credit, total_debit) > 0 else 0
    if large_threshold > 0:
        large_txns = [t for t in transactions
                      if t.get("debit", 0) > large_threshold or t.get("credit", 0) > large_threshold]
        if len(large_txns) > 5:
            anomalies.append(f"{len(large_txns)} transactions exceed 10% of total volume")

    net_flow = total_credit - total_debit

    return {
        "type": "bank_statement",
        "transactions": len(transactions),
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "net_flow": round(net_flow, 2),
        "min_balance": round(min_balance, 2) if min_balance != float("inf") else None,
        "max_balance": round(max_balance, 2) if max_balance != float("-inf") else None,
        "date_range": date_range,
        "anomalies": anomalies,
    }


# ── Balance sheet auditor ────────────────────────────────────────────────

def _audit_balance_sheet(headers: List[str], rows: List[List]) -> Dict:
    total_assets = 0.0
    total_liabilities = 0.0
    total_equity = 0.0
    sections: Dict[str, List[Tuple[str, float]]] = {}
    current_section = "other"

    for row in rows:
        if not row:
            continue
        first = str(row[0]).strip().lower() if row[0] else ""

        if any(t in first for t in ["total", "grand total", "sum"]):
            continue
        if first == "" or len(row) < 2:
            continue

        # Detect section headers
        section_keywords = {
            "assets": ["asset", "property", "plant", "equipment", "investment", "receivable", "cash", "inventory"],
            "liabilities": ["liabilit", "loan", "creditor", "payable", "debt", "borrowing"],
            "equity": ["equity", "capital", "reserve", "retained", "shareholder", "stock"],
        }

        for section, kws in section_keywords.items():
            if any(kw in first for kw in kws):
                current_section = section
                break

        value = _parse_number(row[-1]) if len(row) > 1 else None
        if value is not None:
            label = str(row[0]).strip()
            if current_section not in sections:
                sections[current_section] = []
            sections[current_section].append((label, value))

            if current_section == "assets":
                total_assets += value
            elif current_section == "liabilities":
                total_liabilities += value
            elif current_section == "equity":
                total_equity += value

    anomalies = []
    expected = total_liabilities + total_equity
    if total_assets > 0 and abs(total_assets - expected) / total_assets > 0.05:
        anomalies.append(
            f"Balance sheet mismatch: Assets ({total_assets:.2f}) ≠ Liabilities+Equity ({expected:.2f})"
            f" (diff: {abs(total_assets - expected):.2f})"
        )

    current_ratio = None
    debt_equity = None
    current_assets = sum(v for k, items in sections.items() if "asset" in k for _, v in items)
    current_liabilities = sum(v for k, items in sections.items() if "liabilit" in k for _, v in items)
    if current_liabilities > 0:
        current_ratio = round(current_assets / current_liabilities, 2)
    if total_equity > 0:
        debt_equity = round(total_liabilities / total_equity, 2)

    return {
        "type": "balance_sheet",
        "total_assets": round(total_assets, 2),
        "total_liabilities": round(total_liabilities, 2),
        "total_equity": round(total_equity, 2),
        "current_ratio": current_ratio,
        "debt_equity_ratio": debt_equity,
        "anomalies": anomalies,
        "item_count": sum(len(items) for items in sections.values()),
    }


# ── Profit & Loss auditor ────────────────────────────────────────────────

def _audit_profit_loss(headers: List[str], rows: List[List]) -> Dict:
    total_income = 0.0
    total_expenses = 0.0
    items: Dict[str, List[Tuple[str, float]]] = {}
    current_category = "other"

    for row in rows:
        if not row:
            continue
        first = str(row[0]).strip().lower() if row[0] else ""

        if any(t in first for t in ["total", "grand total"]):
            continue
        if first in ("", "particulars", "description", "amount"):
            continue
        if len(row) < 2:
            continue

        cat_keywords = {
            "income": ["income", "revenue", "sales", "fee", "interest", "commission", "gain"],
            "expense": ["expense", "cost", "salary", "rent", "utility", "depreciation",
                        "maintenance", "professional", "legal", "office", "travel", "insurance"],
        }

        for cat, kws in cat_keywords.items():
            if any(kw in first for kw in kws):
                current_category = cat
                break

        value = _parse_number(row[-1])
        if value is not None:
            label = str(row[0]).strip()
            if current_category not in items:
                items[current_category] = []
            items[current_category].append((label, value))
            if current_category == "income":
                total_income += value
            elif current_category == "expense":
                total_expenses += value

    net_profit = round(total_income - total_expenses, 2)
    profit_margin = round((net_profit / total_income) * 100, 2) if total_income > 0 else None

    anomalies = []
    if profit_margin is not None and profit_margin < 0:
        anomalies.append(f"Net loss: {abs(net_profit):.2f} (margin: {profit_margin}%)")
    if total_income > 0 and total_expenses > total_income * 0.95:
        anomalies.append(f"Expenses consume {round(total_expenses/total_income*100, 1)}% of income — thin margin")

    return {
        "type": "profit_loss",
        "total_income": round(total_income, 2),
        "total_expenses": round(total_expenses, 2),
        "net_profit": net_profit,
        "profit_margin": profit_margin,
        "anomalies": anomalies,
        "item_count": sum(len(items) for items in items.values()),
    }


# ── General financial auditor ────────────────────────────────────────────

def _audit_general(headers: List[str], rows: List[List]) -> Dict:
    numeric_cols = []
    for i, h in enumerate(headers):
        vals = [_parse_number(r[i]) for r in rows[:20] if i < len(r)]
        if any(v is not None for v in vals):
            numeric_cols.append(i)

    summary = []
    for col_idx in numeric_cols:
        vals = [_parse_number(r[col_idx]) for r in rows if col_idx < len(r)]
        vals = [v for v in vals if v is not None]
        if vals:
            col_name = str(headers[col_idx]) if col_idx < len(headers) else f"Col{col_idx}"
            summary.append({
                "column": col_name,
                "sum": round(sum(vals), 2),
                "avg": round(sum(vals) / len(vals), 2),
                "min": round(min(vals), 2),
                "max": round(max(vals), 2),
                "count": len(vals),
            })

    return {
        "type": "general",
        "sheets": 1,
        "rows": len(rows),
        "columns": len(headers),
        "numeric_summary": summary,
        "anomalies": [],
    }


# ── Main entry point ─────────────────────────────────────────────────────

def audit_excel(file_bytes: bytes, filename: str = "report.xlsx") -> str:
    """Analyze an Excel financial file and return a formatted audit report."""
    try:
        import openpyxl
    except ImportError:
        return "❌ openpyxl not installed. Cannot audit Excel files."

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    report_parts = [f"📊 <b>Financial Audit</b>: {filename}\n"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]

        doc_type = _detect_doc_type(headers, data_rows)

        if doc_type == "bank_statement":
            result = _audit_bank_statement(headers, data_rows)
            report_parts.append(_fmt_bank_report(sheet_name, result))
        elif doc_type == "balance_sheet":
            result = _audit_balance_sheet(headers, data_rows)
            report_parts.append(_fmt_bs_report(sheet_name, result))
        elif doc_type == "profit_loss":
            result = _audit_profit_loss(headers, data_rows)
            report_parts.append(_fmt_pl_report(sheet_name, result))
        else:
            result = _audit_general(headers, data_rows)
            report_parts.append(_fmt_general_report(sheet_name, result))

    return "\n".join(report_parts)


# ── Formatting helpers ───────────────────────────────────────────────────

def _fmt_bank_report(sheet: str, r: Dict) -> str:
    lines = [f"━━━ <b>Bank Statement</b>: {sheet} ━━━"]
    if r.get("date_range"):
        lines.append(f"📅 {r['date_range'][0]} → {r['date_range'][1]}")
    lines.extend([
        f"📝 Transactions: <b>{r['transactions']}</b>",
        f"💰 Credits: <b>{r['total_credit']:,.2f}</b>",
        f"💸 Debits:  <b>{r['total_debit']:,.2f}</b>",
        f"📊 Net flow: <b>{r['net_flow']:+,.2f}</b>",
    ])
    if r.get("min_balance") is not None:
        lines.append(f"📉 Min balance: {r['min_balance']:,.2f}")
    if r.get("max_balance") is not None:
        lines.append(f"📈 Max balance: {r['max_balance']:,.2f}")
    if r.get("anomalies"):
        lines.append(f"\n⚠️ <b>{len(r['anomalies'])} anomaly/ies:</b>")
        for a in r["anomalies"]:
            lines.append(f"  ⚠ {a}")
    return "\n".join(lines)


def _fmt_bs_report(sheet: str, r: Dict) -> str:
    lines = [f"━━━ <b>Balance Sheet</b>: {sheet} ━━━"]
    lines.extend([
        f"🏢 Assets:      <b>{r['total_assets']:,.2f}</b>",
        f"📋 Liabilities: <b>{r['total_liabilities']:,.2f}</b>",
        f"👥 Equity:      <b>{r['total_equity']:,.2f}</b>",
    ])
    if r.get("current_ratio") is not None:
        status = "✅" if 1.2 <= r["current_ratio"] <= 3.0 else "⚠️"
        lines.append(f"{status} Current ratio: <b>{r['current_ratio']}</b> (healthy: 1.2–3.0)")
    if r.get("debt_equity_ratio") is not None:
        status = "✅" if r["debt_equity_ratio"] <= 2.0 else "⚠️"
        lines.append(f"{status} Debt/Equity: <b>{r['debt_equity_ratio']}</b> (healthy: ≤2.0)")
    lines.append(f"📦 Items: {r.get('item_count', '?')}")
    if r.get("anomalies"):
        lines.append(f"\n⚠️ <b>{len(r['anomalies'])} anomaly/ies:</b>")
        for a in r["anomalies"]:
            lines.append(f"  ⚠ {a}")
    return "\n".join(lines)


def _fmt_pl_report(sheet: str, r: Dict) -> str:
    lines = [f"━━━ <b>Profit & Loss</b>: {sheet} ━━━"]
    lines.extend([
        f"📈 Income:   <b>{r['total_income']:,.2f}</b>",
        f"📉 Expenses: <b>{r['total_expenses']:,.2f}</b>",
        f"🎯 Net:      <b>{r['net_profit']:+,.2f}</b>",
    ])
    if r.get("profit_margin") is not None:
        status = "✅" if r["profit_margin"] > 0 else "⚠️"
        lines.append(f"{status} Margin: <b>{r['profit_margin']}%</b>")
    lines.append(f"📦 Items: {r.get('item_count', '?')}")
    if r.get("anomalies"):
        lines.append(f"\n⚠️ <b>{len(r['anomalies'])} anomaly/ies:</b>")
        for a in r["anomalies"]:
            lines.append(f"  ⚠ {a}")
    return "\n".join(lines)


def _fmt_general_report(sheet: str, r: Dict) -> str:
    lines = [f"━━━ <b>Financial Sheet</b>: {sheet} ━━━"]
    lines.append(f"📊 {r['rows']} rows × {r['columns']} cols")
    for col in r.get("numeric_summary", []):
        lines.append(f"\n<b>{col['column']}</b>:")
        lines.append(f"  Sum: {col['sum']:,.2f}  Avg: {col['avg']:,.2f}")
        lines.append(f"  Min: {col['min']:,.2f}  Max: {col['max']:,.2f}")
    if r.get("anomalies"):
        for a in r["anomalies"]:
            lines.append(f"⚠ {a}")
    return "\n".join(lines)
